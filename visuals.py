import itertools
import os
from pathlib import Path
import re
import openpyxl
import pandas as pd
import xlsxwriter
import xlwings as xw
import itertools
from . import movements, helper, utils

l = utils.GrabLogs()


def count_names(name_list):
    """
    The function counts the number of names in a given list.

    :param name_list: A list of names that we want to count
    :return: the length of the input list `name_list`.
    """
    return len(name_list)


# The `ShowData` class contains methods for extracting and manipulating data from a dataframe based on
# specific criteria.
class ShowData:
    def __init__(self, df, comb_val) -> None:  # df stands for dataframe
        self.df = df
        self.comb_val = comb_val

    def count_distinct(self):
        """
        This function counts the number of distinct nationalities in a specific home function and level.
        :return: the count of distinct nationalities in the "MB-1" level of the dataframe that match a
        specific region value.
        """
        nat_list = self.show_mb_1()
        uni = []
        for idx, colidx in self.df.iterrows():
            for col in nat_list:
                if (col == colidx["FullName"] or col in colidx["FullName"]) and colidx[
                    "Nationality2"
                ] not in uni:
                    uni.append(colidx["Nationality2"])
        return len(uni)

    def show_mb_1(self):
        """
        This function returns a list of full names of individuals in the "MB-1" level who belong to a
        specific region.
        :return: The function `show_mb_1` returns a list of full names of individuals who belong to the
        "MB-1" level and whose region matches the value of `self.comb_val`.
        """

        name_by_level = self.df.loc[
            (self.df["MBLevel2"] == "MB-1")
            & (self.df["PayGradeCode"].isin(["GR37", "GR38", "GR40", "GR41", "GRMB"]))
            & (self.df["MBFinalInclude2"] == "Include")
        ]
        name_by_reg = []
        for i, col in name_by_level.iterrows():
            if (
                self.comb_val
                in [col["LT_SecondView2"], col["LT_SharedResponsibility2"]]
                and col["FullName"] not in name_by_reg
            ):
                name_by_reg.append(col["FullName"])
        return name_by_reg

    def show_mb_2(self):
        name_by_level = self.df.loc[
            (self.df["MBLevel2"] == "MB-2")
            & (
                self.df["PayGradeCode"].isin(
                    ["GR36", "GR37", "GR38", "GR40", "GR41", "GRMB"]
                )
            )
            & (self.df["MBFinalInclude2"] == "Include")
        ]
        name_by_reg = []
        for i, col in name_by_level.iterrows():
            if (
                self.comb_val
                in [col["LT_SecondView2"], col["LT_SharedResponsibility2"]]
                and col["FullName"] not in name_by_reg
            ):
                name_by_reg.append(col["FullName"])
        return name_by_reg

    def show_mb_1_wgender(self):
        name_by_level = self.df.loc[
            (self.df["MBLevel2"] == "MB-1")
            & (self.df["PayGradeCode"].isin(["GR37", "GR38", "GR40", "GR41", "GRMB"]))
            & (self.df["MBFinalInclude2"] == "Include")
        ]
        name_by_reg = []
        for i, col in name_by_level.iterrows():
            if (
                self.comb_val == col["LT_SecondView2"]
                or self.comb_val == col["LT_SharedResponsibility2"]
                and col["FullName"] not in name_by_reg
            ):
                name_by_reg.append([col["FullName"], col["Gender"]])
        return name_by_reg

    def show_cross_ind(self):
        """
        This function returns a list of employees' information who have "Yes" in the "Experience"
        column and match the specified region or region and area combination.
        :return: a list of lists containing the full name, gender, region, area, and pay grade of
        employees who have "Yes" in the "Experience" column of the "MB-1" level in the dataframe,
        and whose region matches the value of the "comb_val" attribute of the object.
        """
        name_by_level = self.df.loc[
            ((self.df["MBLevel2"] == "MB-1") | (self.df["MBLevel2"] == "MB-2"))
            & (
                self.df["PayGradeCode"].isin(
                    ["GR36", "GR37", "GR38", "GR40", "GR41", "GRMB"]
                )
            )
            & (self.df["Experience"] == "Yes")
            & (self.df["MBFinalInclude2"] == "Include")
        ]
        return [
            [
                col["FullName"],
                col["Gender"],
                col["Region"],
                col["Area"],
                col["PayGradeCode"],
            ]
            for i, col in name_by_level.iterrows()
            if self.comb_val in [col["LT_SecondView2"], col["LT_SharedResponsibility2"]]
        ]

    def get_mb2_population(self):
        """
        This function returns a list of names and genders of individuals in the MB-2 level who belong to
        a specific region.
        :return: The function `get_mb2_population` returns a list of lists containing the full name and
        gender of individuals who belong to the "MB-2" level and whose region matches the `comb_val`
        attribute of the object.
        """
        name_by_level = self.df.loc[
            (self.df["MBLevel2"] == "MB-2")
            & (
                self.df["PayGradeCode"].isin(
                    ["GR36", "GR37", "GR38", "GR40", "GR41", "GRMB"]
                )
            )
            & (self.df["MBFinalInclude2"] == "Include")
        ]
        gen_list = []
        for i, col in name_by_level.iterrows():
            if (
                self.comb_val == col["LT_SecondView2"]
                or self.comb_val == col["LT_SharedResponsibility2"]
                and col["FullName"] not in gen_list
            ):
                gen_list.append([col["FullName"], col["Gender"]])
        return gen_list

    def get_fem_manag(self):
        """
        This function returns a list of female managers' full names by comparing them to a list
        of names and genders from two different levels.
        :return: a list of female managers' full names.
        """
        name_by_level = self.show_mb_1_wgender() + self.get_mb2_population()
        femM_list = []
        for fn_i in name_by_level:
            for _ in range(len(fn_i)):
                if "F" in fn_i[1] and fn_i[0] not in femM_list:
                    femM_list.append(fn_i[0])
        return femM_list

    def show_gr34(self) -> list:
        name_by_level = self.df.loc[
            (self.df["MBFinalInclude2"] == "Include")
            & (
                self.df["PayGradeCode"].isin(
                    [
                        "GR34",
                        "GR35",
                        "GR36",
                        "GR37",
                        "GR38",
                        "GR40",
                        "GR41",
                        "GRMB",
                        "GRMT",
                    ]
                )
            )
            & ((self.df["LT_SecondView2"] == self.comb_val))
            & (self.df["MBSecondView2"] == self.df["MBFinalFullName"]),
            ["FullName", "Gender", "PersonIDExternal"],
        ].drop_duplicates()

        return name_by_level.values.tolist()

    def show_gr34_fem(self):
        """
        This function returns a list of female names from a dataframe based on a combination value and a
        null level.
        :return: a list of female names from a specific region and level.
        """
        name_by_level = self.show_gr34()
        gen_list = []
        for val in name_by_level:
            for _ in range(len(val)):
                if "F" in val[1] and val[0] not in gen_list:
                    gen_list.append(val[0])
        return gen_list

    def get_nationality_names(self):
        """
        This function retrieves a list of names, genders, and nationalities from a dataframe based on a
        specified region and level.
        :return: a list of lists containing the full name, gender, and nationality of individuals who
        have a level of "MB-1" and whose region matches a specified value (stored in the variable
        `self.comb_val`).
        """
        name_by_level = self.df.loc[
            (self.df["MBLevel2"] == "MB-1")
            & (self.df["PayGradeCode"].isin(["GR37", "GR38", "GR40", "GR41", "GRMB"]))
            & (self.df["MBFinalInclude2"] == "Include")
        ]
        nat_list = []
        for i, col in name_by_level.iterrows():
            if (
                self.comb_val == col["LT_SecondView2"]
                or self.comb_val == col["LT_SharedResponsibility2"]
                and col["FullName"] not in nat_list
            ):
                nat_list.append([col["FullName"], col["Gender"], col["Nationality2"]])
        return nat_list

    def get_american_nat(self):
        """
        This function returns a list of full names of American players at a specific level and region.
        :return: a list of full names of American players who belong to the "MB-1" level and have a
        region that matches the value of `self.comb_val`.
        """
        name_by_level = self.df.loc[
            (self.df["MBLevel2"] == "MB-1")
            & (self.df["Nationality2"] == "American")
            & (self.df["PayGradeCode"].isin(["GR37", "GR38", "GR40", "GR41", "GRMB"]))
            & (self.df["MBFinalInclude2"] == "Include")
        ]
        ame_name = []
        for _, col in name_by_level.iterrows():
            if (
                self.comb_val
                in [col["LT_SecondView2"], col["LT_SharedResponsibility2"]]
                and col["FullName"] not in ame_name
            ):
                ame_name.append(col["FullName"])
        return ame_name


# The `FormTable` class inherits from the `ShowData` class and takes a dataframe and combination value
# as input parameters in its constructor.
class FormTable(ShowData):
    def __init__(self, df, comb_val) -> None:
        super().__init__(df, comb_val)

    def gen_mb1(self):
        """
        This function generates a pandas DataFrame containing a list of names with an appended count of
        the number of names in the list.
        :return: a pandas DataFrame containing a list of names from the "mb-1" category, along with a
        counter appended to the end of the list. The DataFrame has one column named "FullName".
        """
        mb1_list = self.show_mb_1()
        mb1_num = count_names(mb1_list)
        mb1_list.append(str(mb1_num))
        return pd.DataFrame(mb1_list, columns=["Full Name"])

    def gen_population(self, q):
        """
        This function generates a population table as numbers, including mb-1/2.
        :return: The function `gen_population` is returning a pandas DataFrame `pop_num_df` containing the
        number of individuals in the population, including mb-1/2. The DataFrame has two rows, one for
        mb-1/2 and one for the total number of individuals in the population. The column name is "Q Actual".
        """
        mb2_num = count_names(ShowData.get_mb2_population(self))
        tot_mb_num = count_names(ShowData.show_mb_1(self)) + mb2_num
        num_list = [mb2_num, tot_mb_num]
        return pd.DataFrame(num_list, columns=[q])

    def gen_corrected_population(self, mb2_df, q):
        mb2_num = get_real_mb2_population(mb2_df, q)
        tot_mb12 = len(self.show_mb_1()) + mb2_num
        num_list = [mb2_num, tot_mb12]

        return pd.DataFrame(num_list, columns=[q])

    def gen_female_srt(self, q):
        """
        Generate a DataFrame containing the number and percentage of females in senior teams
        for a given population.

        Parameters:
            q (str): The label for the quarter or time period.

        Returns:
            DataFrame: A pandas DataFrame containing the count and percentage of females in senior teams
            for the specified population.
        """

        fem_senior_count = count_names(ShowData.get_fem_manag(self))
        total_senior_count = count_names(ShowData.show_mb_1(self)) + count_names(
            ShowData.get_mb2_population(self)
        )

        if total_senior_count > 0:
            fem_senior_percentage = (fem_senior_count / total_senior_count) * 100
        else:
            fem_senior_percentage = 0
        return pd.DataFrame(
            {q: [fem_senior_count, f"{round(fem_senior_percentage)}%"]}, columns=[q]
        )

    def corrected_female_srt(self, mb2_df, q):
        female_srt_count = get_real_female_population(mb2_df, q)
        mb2_num = get_real_mb2_population(mb2_df, q)
        tot_mb12 = len(self.show_mb_1()) + mb2_num

        fem_sen_perc = (female_srt_count / tot_mb12) * 100 if tot_mb12 > 0 else 0
        return pd.DataFrame(
            {q: [female_srt_count, f"{round(fem_sen_perc)}%"]}, columns=[q]
        )

    def gen_gr34(self):
        """
        This function generates a dataframe containing the full name and gender of all personnel on a given LT
        :returns: The function is returning a Dataframe with 'Full name' and 'gender'
        """
        data = ShowData.show_gr34(self)
        return pd.DataFrame(data, columns=["FullName", "Gender", "PersonIDExternal"])

    def gen_fem34(self, q):
        """
        Generate a DataFrame containing the count and percentage of female names in the gr34+ category.

        Parameters:
            q (str): The label for the quarter or time period.

        Returns:
            DataFrame: A pandas DataFrame containing the count of names in the "gr34+" group (including GG),
            the count of female names in the "gr34+" group, and the percentage of female names in the "gr34+" group.
        """

        total_names_count = count_names(ShowData.show_gr34(self))
        fem_names_count = count_names(ShowData.show_gr34_fem(self))
        if total_names_count > 0:
            fem_percentage = (fem_names_count / total_names_count) * 100
        else:
            fem_percentage = 0

        return pd.DataFrame(
            {q: [total_names_count, fem_names_count, f"{round(fem_percentage)}%"]},
            columns=[q],
        )

    def gen_rlft(self):
        """
        Generate a pandas DataFrame containing information about the nationality distribution
        and the number of US nationals in a given dataset.

        Returns:
            DataFrame: A pandas DataFrame containing the required information.
        """
        nationality_data = ShowData.get_nationality_names(self)

        distinct_nat_percentage = (
            (ShowData.count_distinct(self) / count_names(ShowData.show_mb_1(self)))
            * 100
            if count_names(ShowData.show_mb_1(self)) > 0
            else 0
        )

        us_nationals_count = count_names(ShowData.get_american_nat(self))

        additional_data = [
            ["", "", ""],
            ["Distinct nat %", "", f"{str(round(distinct_nat_percentage))}%"],
            ["US Nationals", "", str(us_nationals_count)],
        ]

        data = nationality_data + additional_data

        return pd.DataFrame(
            data, columns=["FullName", "Gender", "Nationality2"]
        ).drop_duplicates()

    def gen_cross_ind(self):
        """
        This function generates a cross industry table from a list and returns it as a pandas dataframe.
        :return: The function `gen_cross_ind` returns a pandas DataFrame object that contains information
        about employees' full name, gender, region, area, and pay grade code.
        """
        # cross industry table
        test_list = self.show_cross_ind()
        return pd.DataFrame(
            test_list,
            columns=["FullName", "Gender", "Region", "Area", "PayGrade Code"],
        ).drop_duplicates()

    def gen_us_nationals(self):
        """
        This function generates a list of American nationals and returns it as a pandas dataframe.
        :return: The function `gen_us_nationals` returns a pandas DataFrame containing a list of American
        nationals' full names. If the list is empty, the function returns a DataFrame with a single row
        containing the string "No US Nationals for this Region".
        """
        name_list = self.get_american_nat()
        name_list_df = pd.DataFrame(name_list, columns=["FullName"])
        if len(name_list_df.index) == 0:
            name_list_df.loc[len(name_list_df)] = "No US Natinals for this Region"
        return name_list_df


def get_real_mb2_population(mb2_df, q_filter):
    df = mb2_df.loc[mb2_df[q_filter] == 1]
    return len(df)


def get_real_female_population(mb2_df, q_filter):
    df = mb2_df.loc[(mb2_df[q_filter] == 1) & (mb2_df["Gender"] == "F")]
    return len(df)


def gen_mb2_population_on_quarters(df, comb_val):
    # sourcery skip: instance-method-first-arg-name
    """
    Generate a DataFrame representing the MB-2 population with snapshot data.

    This function takes a DataFrame 'df' and a combination value 'comb_val' as inputs.
    It processes the data to create a DataFrame containing MB-2 population information
    along with snapshots of data.

    Parameters:
        df (DataFrame): The input DataFrame containing relevant data.
        comb_val (str): The combination value for filtering purposes.

    Returns:
        DataFrame: A DataFrame representing the MB-2 population with snapshot data.
    """
    name_by_level = df.loc[df["MBLevel2"] != "MB-2"].index
    df.drop(name_by_level, inplace=True)
    gen_list = []
    for i, col in df.iterrows():
        if (
            comb_val == col["LT_SecondView2"]
            or comb_val == col["LT_SharedResponsibility2"]
            and col["FullName"] not in gen_list
        ):
            gen_list.append([col["FullName"], col["Gender"]])

    col_list = helper.get_quarter_month_list()
    mb2_pop_df = pd.DataFrame(gen_list, columns=["FullName", "Gender"])
    df_col = [idx for idx, idval in col_list.items()]
    mb2_pop_df = mb2_pop_df.reindex(mb2_pop_df.columns.tolist() + df_col, axis=1)

    name_list = mb2_pop_df["FullName"].tolist()
    wsnap = [
        [xval["FullName"], xval["DateLabel"]]
        for x, xval in df.iterrows()
        if xval["FullName"] in name_list
        and xval["LT_SecondView2"] == comb_val
        or xval["LT_SharedResponsibility2"] == comb_val
    ]
    wsnap = helper.add_quarter_for_snap2list(wsnap)
    mb2_pop_df = mb2_pop_df.fillna(0)
    for idx, ival in mb2_pop_df.iterrows():
        for jval in wsnap:
            if ival["FullName"] in jval and jval[1] in mb2_pop_df.columns:
                mb2_pop_df.at[idx, jval[1]] = 1

    mb2_pop_df = mb2_pop_df.fillna(0)
    mb2_pop_df = mb2_pop_df.drop_duplicates()
    return mb2_pop_df.loc[:, : helper.map_quarters()[0]]


def gen_g34df_last(comb_val):
    """
    Given an LT (in this case a combobox value), this function generates a DataFrame containing
    information from various sources for employees with a specific LT value and GR34 paygrade.
    """
    df = helper.RawDataCleanPrep().split_by_snap()[1]
    adj_df = helper.split_by_snap_custom(
        helper.RawDataCleanPrep().update_df(), "DateLabel"
    )[1]

    filtered_df = df[
        df["PayGradeCode"].isin(
            ["G34", "G35", "G36", "GR37", "GR38", "GR40", "GR41", "GRMT"]
        )
        & (df["MBFinalInclude2"] == "Include")
        # & ((df["LT_SecondView2"] == comb_val))
        & (df["MBFinalFullName"] == df["MBFinalFullName"])
    ]

    merged_df = filtered_df.merge(adj_df, on="PersonIDExternal", how="inner")
    merged_df = merged_df.drop_duplicates()

    merged_df.columns = [col.replace("_y", "") for col in merged_df.columns]

    return merged_df.loc[
        merged_df["MBFinalInclude2"] == "Include", merged_df.columns[26:]
    ]


def write_report(comb_val):
    tot_period = helper.map_q()
    period = f"{tot_period[0]} v {tot_period[1]}"
    filename = f"Comparisson {period} D&I MB Reporting {comb_val}.xlsx"
    fl_loc = f"comp_file/{filename}"
    wb = xlsxwriter.Workbook(Path(fl_loc))
    writer = pd.ExcelWriter(fl_loc, engine="openpyxl", mode="w+")

    # class init
    kprev = FormTable(helper.RawDataCleanPrep().split_by_snap()[0], comb_val)
    kact = FormTable(helper.RawDataCleanPrep().split_by_snap()[1], comb_val)

    # for the main data sheet
    mdws = wb.add_worksheet("Main Data")

    # adding quarter names in dataframe form
    data = [tot_period[0]]
    data1 = [tot_period[1]]
    dfp = pd.DataFrame(data, columns=["Q List"])
    dfa = pd.DataFrame(data1, columns=["Q List"])

    dfp.to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=0,
        startrow=3,
        index=False,
        header=False,
    )
    dfa.to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=2,
        startrow=3,
        index=False,
        header=False,
    )
    kprev.gen_mb1().to_excel(
        writer, sheet_name=mdws.get_name(), startcol=0, startrow=4, index=False
    )
    kact.gen_mb1().to_excel(
        writer, sheet_name=mdws.get_name(), startcol=2, startrow=4, index=False
    )

    kprev.gen_population(tot_period[0]).to_excel(
        writer, sheet_name=mdws.get_name(), startcol=5, startrow=3, index=False
    )
    kact.gen_population(tot_period[1]).to_excel(
        writer, sheet_name=mdws.get_name(), startcol=6, startrow=3, index=False
    )

    kprev.gen_female_srt(tot_period[0]).to_excel(
        writer, sheet_name=mdws.get_name(), startcol=5, startrow=10, index=False
    )
    kact.gen_female_srt(tot_period[1]).to_excel(
        writer, sheet_name=mdws.get_name(), startcol=6, startrow=10, index=False
    )

    kprev.gen_fem34(tot_period[0]).to_excel(
        writer, sheet_name=mdws.get_name(), startcol=5, startrow=16, index=False
    )
    kact.gen_fem34(tot_period[1]).to_excel(
        writer, sheet_name=mdws.get_name(), startcol=6, startrow=16, index=False
    )

    r = max(kact.gen_mb1().shape[0] + 6, kprev.gen_mb1().shape[0] + 6, 23) + 2

    dfp.to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=0,
        startrow=r + 1,
        index=False,
        header=False,
    )

    dfa.to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=4,
        startrow=r + 1,
        index=False,
        header=False,
    )

    kprev.gen_rlft().to_excel(
        writer, sheet_name=mdws.get_name(), startcol=0, startrow=r + 2, index=False
    )
    kact.gen_rlft().to_excel(
        writer, sheet_name=mdws.get_name(), startcol=4, startrow=r + 2, index=False
    )

    row = max(kprev.gen_rlft().shape[0] + r + 5, kact.gen_rlft().shape[0] + r + 5)

    dfp.to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=0,
        startrow=row + 1,
        index=False,
        header=False,
    )
    dfa.to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=0,
        startrow=row + 1 + kprev.gen_cross_ind().shape[0] + 4,
        index=False,
        header=False,
    )
    kprev.gen_cross_ind().to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=0,
        startrow=row + 2,
        index=False,
    )
    kact.gen_cross_ind().to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=0,
        startrow=row + kprev.gen_cross_ind().shape[0] + 6,
        index=False,
    )

    # MB-2 name list
    worksheet2 = wb.add_worksheet("MB-2 Name List")

    gen_mb2_population_on_quarters(helper.assign_lt(), comb_val).to_excel(
        writer,
        sheet_name=worksheet2.get_name(),
        startcol=1,
        startrow=5,
        index=False,
    )

    # ws5 data
    worksheet5 = wb.add_worksheet("GR34 Full Name")
    gen_g34df_last(comb_val).to_excel(
        writer, sheet_name=worksheet5.get_name(), startcol=1, startrow=3, index=False
    )
    # ws3 data
    worksheet3 = wb.add_worksheet("Movements")
    added_list = movements.filter_added(comb_val)
    removed_list = movements.filter_removed(comb_val)
    added_list.to_excel(
        writer, sheet_name=worksheet3.get_name(), startcol=1, startrow=3, index=False
    )
    removed_list.to_excel(
        writer, sheet_name=worksheet3.get_name(), startcol=6, startrow=3, index=False
    )

    wb.close()
    writer.close()

    # adding details on each worksheet
    from openpyxl.styles import Alignment, Border, Side

    def set_border(ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    dest_path = os.path.join(os.getcwd(), fl_loc).replace("\\", "\\")
    try:
        workb = openpyxl.load_workbook(dest_path)

        # adding details to main data sheet
        worksheet1 = workb["Main Data"]
        worksheet1["A2"] = "MB-1 total population"
        worksheet1.merge_cells("A2:C2")
        cc = worksheet1["A2"]
        cc.alignment = Alignment(horizontal="center")

        set_border(worksheet1, "A4:A4")
        set_border(worksheet1, "C4:C4")
        set_border(worksheet1, "E4:G6")
        set_border(worksheet1, "E11:G13")
        set_border(worksheet1, "E17:G20")

        worksheet1["E2"] = "MB-2 Total Population"
        worksheet1.merge_cells("E2:G2")
        ccr = worksheet1["E2"]
        ccr.alignment = Alignment(horizontal="center")

        worksheet1["E5"] = "Total MB-2 population"
        worksheet1["E5"].alignment = Alignment(horizontal="center")

        worksheet1["E6"] = "Total MB-1/2 population"
        worksheet1["E6"].alignment = Alignment(horizontal="center")

        worksheet1["E9"] = "Female % in senior teams"
        worksheet1.merge_cells("E9:G9")
        worksheet1["E9"].alignment = Alignment(horizontal="center")

        worksheet1["E12"] = "MB-1/2 Female #"
        worksheet1["E12"].alignment = Alignment(horizontal="center")

        worksheet1["E13"] = "MB-1/2 Female %"
        worksheet1["E13"].alignment = Alignment(horizontal="center")

        worksheet1["E15"] = "Female % G34+ incl. GGs"
        worksheet1.merge_cells("E15:G15")
        worksheet1["E15"].alignment = Alignment(horizontal="center")

        worksheet1["E18"] = "Total Population"
        worksheet1["E18"].alignment = Alignment(horizontal="center")

        worksheet1["E19"] = "Female #"
        worksheet1["E19"].alignment = Alignment(horizontal="center")

        worksheet1["E20"] = "Female %"
        worksheet1["E20"].alignment = Alignment(horizontal="center")

        kact_shape = kact.gen_mb1().shape[0] + 4
        kprev_shape = kprev.gen_mb1().shape[0] + 4

        row_rlft_ref = max(r, kact_shape, kprev_shape)

        worksheet1[f"A{row_rlft_ref}"] = "50% distinct nationalities in RLT/FLT"
        worksheet1.merge_cells(f"A{row_rlft_ref}:G{row_rlft_ref}")
        worksheet1[f"A{row_rlft_ref}"].alignment = Alignment(horizontal="center")

        title_cell_value = row if row_rlft_ref <= 22 else row
        title_cell = f"A{title_cell_value}"
        title_cell_range = f"A{title_cell_value}:G{title_cell_value}"

        worksheet1[title_cell] = "Cross Industry hires in senior teams past 3Y"
        worksheet1.merge_cells(title_cell_range)
        worksheet1[title_cell].alignment = Alignment(horizontal="center")
        try:
            worksheet1[f"C{row+2}"] = count_names(kprev.show_cross_ind())
            worksheet1[f"C{row+2}"].alignment = Alignment(horizontal="center")
        except AttributeError:
            worksheet1[f"C{row+2}"] = count_names(kprev.show_cross_ind())
            worksheet1[f"C{row+2}"].alignment = Alignment(horizontal="center")
        try:
            perc = round(
                (
                    count_names(kprev.show_cross_ind())
                    / (
                        count_names(kprev.show_mb_1())
                        + count_names(kprev.get_mb2_population())
                    )
                    * 100
                ),
            )
            worksheet1[f"D{row+2}"] = f"{perc}%"
            worksheet1[f"D{row+2}"].alignment = Alignment(horizontal="center")
        except ZeroDivisionError:
            perc = 0
            worksheet1[f"D{row+2}"] = f"{perc}%"
            worksheet1[f"D{row+2}"].alignment = Alignment(horizontal="center")

        worksheet1[f"C{row + kprev.gen_cross_ind().shape[0] +6}"] = count_names(
            kact.show_cross_ind()
        )
        worksheet1[f"C{row + kprev.gen_cross_ind().shape[0] +6}"].alignment = Alignment(
            horizontal="center"
        )

        try:
            ci_pos = row + kprev.gen_cross_ind().shape[0] + 6
            ci_proc = round(
                (
                    count_names(kact.show_cross_ind())
                    / (
                        count_names(kact.show_mb_1())
                        + count_names(kact.get_mb2_population())
                    )
                    * 100
                )
            )
        except ZeroDivisionError:
            ci_proc = "0%"
        worksheet1[f"D{ci_pos}"] = f"{ci_proc}%"
        worksheet1[f"D{ci_pos}"].alignment = Alignment(horizontal="center")

        # adding details to MB-2 Name List sheet
        try:
            worksheet2 = workb["MB-2 Name List"]
            worksheet2["B3"] = "MB-2 Name List"
            worksheet2.merge_cells("B3:K3")
            worksheet2["B3"].alignment = Alignment(horizontal="center")
            set_border(
                worksheet2,
                f"B6:AA{6 + gen_mb2_population_on_quarters( helper.assign_lt(), comb_val).shape[0]}",
            )
        except Exception:
            l.form_log(Exception, 30)

        # adding next sheets
        worksheet3 = workb["Movements"]
        worksheet3["B2"] = "Movement List"
        worksheet3["B2"].alignment = Alignment(horizontal="center")

        worksheet4 = workb.create_sheet("Comments")
        worksheet4[
            "B3"
        ] = "Please write the additional comments here, following the formula:"
        worksheet4["B5"] = "Name"
        worksheet4["C5"] = "Comments"
        worksheet4["A6"] = "Example"
        worksheet4["B6"] = "John Doe"
        worksheet4["C6"] = "New hired on 1st July"

        worksheet4["B3"].alignment = Alignment(horizontal="center")
        worksheet4["B5"].alignment = Alignment(horizontal="center")
        worksheet4["C5"].alignment = Alignment(horizontal="center")

        set_border(worksheet4, "B5:C6")

        worksheet5 = workb["GR34 Full Name"]
        worksheet5["B2"] = "GR34+ Name List"
        worksheet5["B2"].alignment = Alignment(horizontal="center")

        set_border(worksheet5, f"B4:AE{4+gen_g34df_last(comb_val).shape[0]}")

        workb.save(dest_path)
    except IOError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Main Data"
        wb.save(dest_path)
    l.form_log(f"{fl_loc} has been generated", 20)
    return dest_path


def add_details(file_loc):
    with xw.App() as writer:
        workb = writer.books.open(file_loc)
        # workb.app.screen_updating = False

        workb.sheets["Main Data"].autofit()
        # will autofit columns and rows in sheet. To customize: pass 'c' for columns and 'r' for rows
        worksheet1 = workb.sheets["Main Data"]
        header_text_look(worksheet1, "A2")
        header_text_look(worksheet1, "E2")
        header_text_look(worksheet1, "E9")
        header_text_look(worksheet1, "E15")

        header2_text_look(worksheet1, "A4")
        header2_text_look(worksheet1, "C4")
        header2_text_look(worksheet1, "F4")
        header2_text_look(worksheet1, "G4")
        header2_text_look(worksheet1, "F11")
        header2_text_look(worksheet1, "G11")
        header2_text_look(worksheet1, "F17")
        header2_text_look(worksheet1, "G17")

        header3_text_look(worksheet1, "A5")
        header3_text_look(worksheet1, "C5")

        rlftpos = []
        crossind_pos = []
        for row, col in itertools.product(range(1, 100), range(1, 7)):
            if (
                worksheet1.range((row, col)).value
                == "50% distinct nationalities in RLT/FLT"
            ):
                rlftpos.extend((col, row))
            if (
                worksheet1.range((row, col)).value
                == "Cross Industry hires in senior teams past 3Y"
            ):
                crossind_pos.extend((col, row))

        adaptive_header1_style(worksheet1, rlftpos)
        adaptive_header1_style(worksheet1, crossind_pos)

        adaptive_header2_style(worksheet1, [rlftpos[0], rlftpos[1] + 2])
        adaptive_header2_style(worksheet1, [rlftpos[0] + 4, rlftpos[1] + 2])

        adaptive_q_cell_style(worksheet1, rlftpos)
        adaptive_header3_style(worksheet1, [rlftpos[0] + 4, rlftpos[1] + 3])
        adaptive_header3_style(worksheet1, [rlftpos[0] + 5, rlftpos[1] + 3])
        adaptive_header3_style(worksheet1, [rlftpos[0] + 6, rlftpos[1] + 3])

        adaptive_header2_style(worksheet1, [crossind_pos[0], crossind_pos[1] + 2])
        ci_count_style(worksheet1, [crossind_pos[0] + 2, crossind_pos[1] + 2])

        adaptive_cross_industry_headers(worksheet1, crossind_pos, 3)
        ciq_cell = []
        for r, c in itertools.product(range(crossind_pos[1] + 3, 100), range(1, 7)):
            if re.match("^Q[1-4].[1-5][1-5]", str(worksheet1.range((r, c)).value)):
                ciq_cell.extend((c, r))

        adaptive_header2_style(worksheet1, ciq_cell)
        ci_count_style(worksheet1, [ciq_cell[0] + 2, ciq_cell[1]])

        adaptive_cross_industry_headers(worksheet1, ciq_cell, 1)

        # worksheet2
        worksheet2 = workb.sheets["MB-2 Name List"]
        workb.sheets["MB-2 Name List"].autofit()
        worksheet2["B3"].color = (255, 192, 0)
        worksheet2.range("B3:B3").api.Font.Bold = True
        worksheet2.range("B3:B3").api.Font.ColorIndex = 0

        worksheet2["B6:AA6"].color = (0, 32, 96)
        worksheet2.range("B6:AA6").api.Font.Bold = True
        worksheet2.range("B6:AA6").api.Font.ColorIndex = 2

        worksheet2.range("B3:B3").api.Font.Bold = True
        worksheet2.range("B3:B3").api.Font.ColorIndex = 25

        worksheet3 = workb.sheets["Movements"]
        workb.sheets["Movements"].autofit()

        worksheet3["B2"].color = (255, 192, 0)
        worksheet3.range("B2:B2").api.Font.Bold = True
        worksheet3.range("B2:B2").api.Font.ColorIndex = 25

        worksheet3["B6:E6"].color = (0, 32, 96)
        worksheet3.range("B6:E6").api.Font.Bold = True
        worksheet3.range("B6:E6").api.Font.ColorIndex = 2

        worksheet3["G6:J6"].color = (0, 32, 96)
        worksheet3.range("G6:J6").api.Font.Bold = True
        worksheet3.range("G6:J6").api.Font.ColorIndex = 2

        workb.sheets["Comments"].autofit()
        worksheet4 = workb.sheets["Comments"]

        worksheet4["B3:B3"].color = (255, 192, 0)
        worksheet4.range("B3:B3").api.Font.Bold = True
        worksheet4.range("B3:B3").api.Font.ColorIndex = 25

        worksheet4["B5:C5"].color = (0, 32, 96)
        worksheet4.range("B5:C5").api.Font.Bold = True
        worksheet4.range("B5:C5").api.Font.ColorIndex = 2

        workb.sheets["GR34 Full Name"].autofit()
        worksheet5 = workb.sheets["GR34 Full Name"]

        worksheet5["B2:B2"].color = (255, 192, 0)
        worksheet5.range("B2:B2").api.Font.Bold = True
        worksheet5.range("B2:B2").api.Font.ColorIndex = 25

        worksheet5["B4:AE4"].color = (47, 117, 181)
        worksheet5.range("B4:AE4").api.Font.Bold = True
        worksheet5.range("B4:AE4").api.Font.ColorIndex = 2

        workb.save(file_loc)
    return file_loc


def adaptive_cross_industry_headers(worksheet1, arg1, arg2):
    adaptive_ci_header_style(worksheet1, [arg1[0], arg1[1] + arg2])
    adaptive_ci_header_style(worksheet1, [arg1[0] + 1, arg1[1] + arg2])
    adaptive_ci_header_style(worksheet1, [arg1[0] + 2, arg1[1] + arg2])
    adaptive_ci_header_style(worksheet1, [arg1[0] + 3, arg1[1] + arg2])
    adaptive_ci_header_style(worksheet1, [arg1[0] + 4, arg1[1] + arg2])


def adaptive_q_cell_style(worksheet1, arg1):
    adaptive_header3_style(worksheet1, [arg1[0], arg1[1] + 3])
    adaptive_header3_style(worksheet1, [arg1[0] + 1, arg1[1] + 3])
    adaptive_header3_style(worksheet1, [arg1[0] + 2, arg1[1] + 3])


def adaptive_header1_style(worksheet1, arg1):
    worksheet1[arg1[1] - 1, arg1[0] - 1].color = (112, 48, 160)
    worksheet1[arg1[1] - 1, arg1[0] - 1].api.Font.Bold = True
    worksheet1[arg1[1] - 1, arg1[0] - 1].api.Font.ColorIndex = 2


def ci_count_style(worksheet1, arg1):
    worksheet1[arg1[1] - 1, arg1[0] - 1].color = (0, 112, 192)
    worksheet1[arg1[1] - 1, arg1[0] - 1].api.Font.Bold = True
    worksheet1[arg1[1] - 1, arg1[0] - 1].api.Font.ColorIndex = 2


def adaptive_header2_style(worksheet1, arg_list):
    worksheet1[arg_list[1] - 1, arg_list[0] - 1].color = (47, 117, 181)
    worksheet1[arg_list[1] - 1, arg_list[0] - 1].api.Font.Bold = True
    worksheet1[arg_list[1] - 1, arg_list[0] - 1].api.Font.ColorIndex = 2


def adaptive_header3_style(worksheet1, arg_list):
    worksheet1[arg_list[1] - 1, arg_list[0] - 1].color = (0, 176, 80)
    worksheet1[arg_list[1] - 1, arg_list[0] - 1].api.Font.Bold = True
    worksheet1[arg_list[1] - 1, arg_list[0] - 1].api.Font.ColorIndex = 2


def adaptive_ci_header_style(worksheet1, arg_list):
    worksheet1[arg_list[1] - 1, arg_list[0] - 1].color = (68, 84, 106)
    worksheet1[arg_list[1] - 1, arg_list[0] - 1].api.Font.Bold = True
    worksheet1[arg_list[1] - 1, arg_list[0] - 1].api.Font.ColorIndex = 2


def header_text_look(worksheet1, arg1):
    worksheet1[arg1].color = (112, 48, 160)
    worksheet1.range(arg1).api.Font.Bold = True
    worksheet1.range(arg1).api.Font.ColorIndex = 2


def header2_text_look(worksheet1, arg1):
    worksheet1[arg1].color = (47, 117, 181)
    worksheet1.range(arg1).api.Font.Bold = True
    worksheet1.range(arg1).api.Font.ColorIndex = 2


def header3_text_look(worksheet1, arg1):
    worksheet1[arg1].color = (0, 176, 80)
    worksheet1.range(arg1).api.Font.Bold = True
    worksheet1.range(arg1).api.Font.ColorIndex = 2
