"""
This section of the program will be dedicated to each fixture we need to implement.
It aims to replace a work-in-progress (WIP) process, which involves tasks like 
rectifying raw data errors or enhancing the logical structure of the management board 
(MB) for improved coherence.
"""
from pathlib import Path
from tkinter import messagebox
from tkinter.filedialog import askopenfilename
import pandas as pd
from . import helper, visuals, movements, utils
import xlsxwriter
import os
import openpyxl
import xlwings as xw
import itertools
import re


l = utils.GrabLogs()


def correct_q_data():
    """
    Loads and returns data from an Excel file with specified sheets.

    This function prompts the user to select an Excel file containing data and returns two DataFrames.

    Returns:
        tuple: A tuple containing two DataFrames.

        - The first DataFrame (index 0) contains data from the "data" sheet.
        - The second DataFrame (index 1) contains data from the "MB2 Name List" sheet.

    Notes:
        - The user is prompted to select an Excel file using a file dialog.
        - The Excel file should have two sheets named "data" and "MB2 Name List".
    """
    messagebox.showinfo(
        "Pick a file",
        "Please select the corrected file for the last quarter, in order to continue",
    )
    file = askopenfilename(
        filetypes=[
            ("Excel Files", "*.xlsx"),
        ]
    )

    return pd.read_excel(file, sheet_name="data"), pd.read_excel(
        file,
        sheet_name="MB2 Name List",
    )


@utils.onedrive_operations_decorator
def fix_generate_file(c):
    """
    This function generates a file based on a given parameter and prompts the user to open it.

    :param c: The parameter `c` is a string representing the name of a file to be generated
    :return: the output of the `subprocess.call()` function, which is the return code of the executed
    command.
    """
    try:
        file_gen = add_details(fix_report(c))
    except Exception:
        if kill_prompt := messagebox.askokcancel(
            title="Excel about to close",
            message="This operation cannot be performed because the file is already opened."
            "\n If you press 'OK' to continue the process, all opened Excel files will be closed.",
        ):
            os.system("taskkill /T /IM EXCEL.exe")
            file_gen = add_details(fix_report(c))
            print(f"File for {c} has been generated at {file_gen}")
    return c, file_gen


def get_excel_column_letter(index):
    quotient = (index - 1) // 26
    remainder = (index - 1) % 26
    if quotient > 0:
        return f"{chr(65 + quotient - 1)}{chr(65 + remainder)}"
    else:
        return f"{chr(65 + remainder)}"


class ShowDataAdj:
    def __init__(self, df, comb_val) -> None:
        """
        Initialize an instance of the ShowDataAdj class.

        Parameters:
            df (DataFrame): The main DataFrame containing the data to be manipulated.
            comb_val (str): The value used for combination or filtering purposes.
        """
        self.df = df
        self.comb_val = comb_val

    def get_mb12_team(self):
        return self.df.loc[
            (self.df["Team"] == self.comb_val)
            & (self.df["Level"].isin(["MB-1", "MB-2"]))
            & (self.df["Grade"].isin(["GR36", "GR37", "GR38", "GR40", "GR41", "GRMB"])),
            "Name",
        ]

    def get_mb1_ini(self):
        """
        Get a DataFrame containing individuals from 'self.df' who belong to the specified
        team ('comb_val') and have a level value of 'MB-1'.

        Returns:
            DataFrame: A DataFrame containing the individuals who meet the specified
            team and level conditions.
        """
        return self.df.loc[
            (self.df["Team"] == self.comb_val)
            & (self.df["Level"] == "MB-1")
            & (self.df["Grade"].isin(["GR36", "GR37", "GR38", "GR40", "GR41", "GRMB"])),
            "Name",
        ]

    def get_mb1(self):
        """
        Get a DataFrame containing individuals from 'self.df' who belong to the specified
        team ('comb_val') and have a level value of 'MB-1'.

        Returns:
            DataFrame: A DataFrame containing the individuals who meet the specified
            team and level conditions.
        """
        mb1 = self.df.loc[
            (self.df["Team"] == self.comb_val)
            & (self.df["Level"] == "MB-1")
            & (self.df["Grade"].isin(["GR36", "GR37", "GR38", "GR40", "GR41", "GRMB"])),
            "Name",
        ]
        mb1num = len(mb1)
        mb1_ls = mb1.tolist()
        mb1_ls.append(str(mb1num))
        return pd.DataFrame(mb1_ls, columns=["Full Name"])

    def gen_cross_ind(self, supp_df):
        """
        Get a DataFrame containing individuals from 'supp_df' who belong to the specified
        team ('comb_val'), have levels 'MB-1' or 'MB-2' in 'self.df', and have a value of
        'Yes' for 'Cross Industry' in 'supp_df'.

        Parameters:
            supp_df (DataFrame): The supplementary DataFrame containing individual data.

        Returns:
            DataFrame: A DataFrame containing the individuals who meet the specified
            team, level, and 'Cross Industry' conditions, including columns 'Full Name',
            'Gender', 'Region', 'Area', and 'Pay Grade'.
        """
        selected_names = self.df.loc[
            (self.df["Team"] == self.comb_val)
            & (self.df["Level"].isin(["MB-1", "MB-2"]))
            & (self.df["Grade"].isin(["GR36", "GR37", "GR38", "GR40", "GR41", "GRMB"]))
            & (self.df["Cross Industry"] == "Yes"),
            "Name",
        ]
        supp_df = supp_df.loc[supp_df["MBFinalInclude2"] == "Include"]
        mask = supp_df["FullName"].isin(selected_names)
        result_df = supp_df[mask].copy()

        result_df.dropna(subset=["FullName"], inplace=True)
        result_df = result_df[["FullName", "Gender", "Region", "Area", "PayGradeCode"]]

        return result_df.drop_duplicates()

    def get_mb2(self):
        """
        Get a DataFrame containing individuals from 'self.df' who belong to the specified
        team ('comb_val') and have a level value of 'MB-2'.

        Returns:
            DataFrame: A DataFrame containing the individuals who meet the specified
            team and level conditions.
        """
        return self.df.loc[
            (self.df["Team"] == self.comb_val)
            & (self.df["Level"] == "MB-2")
            & (self.df["Grade"].isin(["GR36", "GR37", "GR38", "GR40", "GR41", "GRMB"]))
        ]

    def get_fem_manag(self):
        """
        Get a DataFrame containing individuals from 'self.df' who belong to the specified
        team ('comb_val') and have a gender value of 'F' (female).

        Returns:
            DataFrame: A DataFrame containing the individuals who meet the specified
            team and gender conditions.
        """
        return self.df.loc[
            (self.df["Team"] == self.comb_val)
            & (self.df["Gender"] == "F")
            & (self.df["Grade"].isin(["GR36", "GR37", "GR38", "GR40", "GR41", "GRMB"]))
        ]

    def get_mb1_nationality(self, supp_df):
        """
        Get a DataFrame containing the names, genders, and nationalities of individuals
        from 'supp_df' who belong to the 'MB-1' level and have their names present in
        the 'self.df' DataFrame based on the specified conditions.

        Parameters:
            supp_df (DataFrame): The supplementary DataFrame containing individual data.

        Returns:
        DataFrame: A DataFrame containing 'Full Name', 'Gender', and 'Nationality'
        columns of individuals from 'supp_df' who meet the conditions.
        """

        # Define the conditions to filter 'self.df' efficiently
        mb1_conditions = (
            (self.df["Team"] == self.comb_val)
            & (self.df["Level"] == "MB-1")
            & (self.df["Grade"].isin(["GR36", "GR37", "GR38", "GR40", "GR41", "GRMB"]))
        )

        names = self.df.loc[mb1_conditions, "Name"].tolist()

        supp_full_names = supp_df.loc[supp_df["MBFinalInclude2"] == "Include"]
        ret = [
            [val["FullName"], val["Gender"], val["Nationality2"]]
            for _, val in supp_full_names.iterrows()
            if val["FullName"] in names
        ]

        distinct_nationalities = len(self.get_distinct_nationality(supp_df))
        total_mb1_count = len(self.get_mb1())
        distinct_nationality_percentage = (
            distinct_nationalities / (total_mb1_count - 1) * 100
            if total_mb1_count > 1
            else 0
        )

        return pd.DataFrame(
            ret
            + [
                ["", "", ""],
                [
                    "Distinct nat %",
                    "",
                    f"{str(round(distinct_nationality_percentage))}%",
                ],
                ["US Nationals", "", str(len(self.get_ame_nat(supp_df)))],
            ],
            columns=["Full Name", "Gender", "Nationality"],
        ).drop_duplicates()

    def get_distinct_nationality(self, supp_df):
        """
        Get an array of distinct nationalities of individuals from 'supp_df'
        who belong to the 'MB-1' level and meet the conditions specified
        in 'get_mb1_nationality' function.

        Parameters:
            supp_df (DataFrame): The supplementary DataFrame containing individual data.

        Returns:
            ndarray: An array containing the unique nationalities of individuals
            who meet the conditions specified in 'get_mb1_nationality'.
        """
        base = self.df.loc[
            (self.df["Team"] == self.comb_val)
            & (self.df["Level"] == "MB-1")
            & (self.df["Grade"].isin(["GR36", "GR37", "GR38", "GR40", "GR41", "GRMB"])),
            "Name",
        ].tolist()

        ret = [
            [val["FullName"], val["Gender"], val["Nationality2"]]
            for _, val in supp_df.iterrows()
            if val["FullName"] in base
            and val["MBFinalInclude2"] == "Include"
            and (
                val["LT_SecondView2"] == self.comb_val
                or val["LT_SharedResponsibility2"] == self.comb_val
            )
        ]
        return pd.DataFrame(
            ret, columns=["Name", "Gender", "Nationality"]
        ).drop_duplicates()

    def get_ame_nat(self, supp_df):
        """
        Get a DataFrame containing the American nationals from 'supp_df' based on specified conditions.

        Parameters:
            supp_df (DataFrame): The supplementary DataFrame containing individual data.

        Returns:
            DataFrame: A DataFrame containing 'Full Name', 'Gender', and 'Nationality' columns
            of American nationals from 'supp_df' who meet the conditions.
        """
        base = self.df.loc[
            (self.df["Team"] == self.comb_val)
            & (self.df["Level"] == "MB-1")
            & (self.df["Grade"].isin(["GR36", "GR37", "GR38", "GR40", "GR41", "GRMB"])),
            ["Name", "Gender", "Nationality"],
        ]
        base_names = base["Name"].tolist()

        supp_df_filtered = supp_df[supp_df["FullName"].isin(base_names)]

        american_nationals = supp_df_filtered[
            supp_df_filtered["Nationality2"] == "American"
        ]

        return american_nationals[["FullName", "Gender", "Nationality2"]]

    def get_population_name_list(self):
        return self.df.loc[self.df["Team"] == self.comb_val]

    def get_population_numbers(self, q):
        total_population = len(self.get_population_name_list())
        mb2pop = len(self.get_mb2())
        return pd.DataFrame([mb2pop, total_population], columns=[q])

    def get_fem_srt(self, q):
        """
        Generate a pandas DataFrame containing the count and percentage of female managers
        for a given population.

        Parameters:
            q (str): The label for the quarter or time period.

        Returns:
            DataFrame: A pandas DataFrame containing the count and percentage of female managers.
        """
        fem_managers_count = len(self.get_fem_manag())
        total_population_count = len(self.get_population_name_list())

        if total_population_count == 0:
            fem_percentage = 0
        else:
            fem_percentage = (fem_managers_count / total_population_count) * 100

        fem_sen_data = [fem_managers_count, f"{round(fem_percentage)}%"]

        return pd.DataFrame(fem_sen_data, columns=[q])


def fix_mb2_moves_per_quarters(history_df, new_df, lt_val) -> pd.DataFrame:
    data = history_df.loc[history_df["Team"] == lt_val, history_df.columns[1:]].fillna(
        0
    )

    new_df = new_df.loc[
        (new_df["MBLevel2"] == "MB-2")
        & new_df["PayGradeCode"].isin(["GR36", "GR37", "GR38", "GR40", "GR41", "GRMB"])
        & (new_df["MBFinalInclude2"] == "Include")
        & (
            (new_df["LT_SecondView2"] == lt_val)
            | (new_df["LT_SharedResponsibility2"] == lt_val)
        ),
        ["FullName", "Gender", "DateLabel"],
    ]
    if not new_df.empty:
        dv_label = new_df["DateLabel"].iat[0]
        q_label = helper.get_quarter_for_snap(dv_label)

        name_set = set(zip(new_df["FullName"], new_df["Gender"]))

        data[q_label] = 0

        for idx, row in data.iterrows():
            if (row["Name"], row["Gender"]) in name_set:
                data.at[idx, q_label] = 1

        new_rows = []
        for name, gender in name_set:
            if not data[(data["Name"] == name) & (data["Gender"] == gender)].empty:
                continue  # Name is already in the DataFrame
            new_row = {"Name": name, "Gender": gender, q_label: 1}
            for col in data.columns:
                if col not in ["Name", "Gender", q_label]:
                    new_row[col] = 0
            new_rows.append(new_row)

        if new_rows:
            data = pd.concat([data] + [pd.DataFrame(new_rows)], ignore_index=True)

        return data


def find_lm_merge(support, cval):
    """
    Find and merge relevant data from the 'support' DataFrame based on a given 'cval' (LT_SecondView2 code).

    Parameters:
    support (DataFrame): The input DataFrame containing support data.
    cval (str): The LT_SecondView2 code to filter and merge data.

    Returns:
    DataFrame: A DataFrame containing merged data for employees and managers with the specified 'cval' in their LT_SecondView2.

    The function performs the following steps:
    1. Filters 'support' DataFrame to include only rows with 'MBFinalInclude2' equal to "Include" and matching 'cval' in 'LT_SecondView2' or 'LT_SharedResponsibility2'.
    2. Separates the filtered data into two DataFrames: 'emp_side' for employees and 'mgr_side' for managers.
    3. Merges 'emp_side' and 'mgr_side' DataFrames based on 'ParentPositionCode' and 'PositionCode'.
    4. Returns a DataFrame containing merged data for managers whose 'LT_SecondView2_mgr' is equal to 'cval'
       and whose 'MBLevel2_emp' is not in ["MB-1", "MB-2", "MB"].

    If an exception occurs during the operation, the function returns the column names of the merged DataFrame.
    """
    emp_side = support.loc[
        (support["MBFinalInclude2"] == "Include")
        & (
            (support["LT_SecondView2"] == cval)
            | (support["LT_SharedResponsibility2"] == cval)
        ),
        [
            "FullName",
            "PositionCode",
            "ParentPositionCode",
            "Gender",
            "MBLevel2",
            "MBSecondView2",
            "MBSharedResponsibility2",
            "PayGradeCode",
            "LineManagerFullName",
            "LineManagerPositionCode",
            "MBFinalInclude2",
            "LT_SecondView2",
            "LT_SharedResponsibility2",
        ],
    ]
    mgr_side = support.loc[
        (
            (support["LT_SecondView2"] == cval)
            | (support["LT_SharedResponsibility2"] == cval)
        ),
        [
            "FullName",
            "PositionCode",
            "ParentPositionCode",
            "Gender",
            "MBLevel2",
            "MBSecondView2",
            "MBSharedResponsibility2",
            "PayGradeCode",
            "LineManagerFullName",
            "LineManagerPositionCode",
            "MBFinalInclude2",
            "LT_SecondView2",
            "LT_SharedResponsibility2",
        ],
    ]

    merged = emp_side.merge(
        mgr_side,
        left_on="ParentPositionCode",
        right_on="PositionCode",
        suffixes=("_emp", "_mgr"),
    )
    try:
        return merged.loc[
            (
                (merged["LT Functioal_emp"] == cval)
                | (merged["LT_SharedResponsibility2_emp"] == cval)
            ),
        ]
    except Exception:
        return merged.columns


def summarize_mb2(history, support, lt_val):
    mov = fix_mb2_moves_per_quarters(history, support, lt_val)
    col_list = mov.columns.tolist()
    new_col_list = col_list[2:]

    sums = [[mov[x].sum()] for x in new_col_list]

    return pd.DataFrame(sums).T


def fix_report(comb_val):
    tot_period = helper.RawDataCleanPrep().map_q()

    period = f"{tot_period[0]} v {tot_period[1]}"
    filename = f"Comparisson {period} D&I MB Reporting {comb_val}.xlsx"
    fl_loc = f"comp_file/{filename}"

    wb = xlsxwriter.Workbook(Path(fl_loc))
    writer = pd.ExcelWriter(fl_loc, engine="openpyxl", mode="w+")

    # misc tools and declarations

    correction = correct_q_data()
    prevq_correction = correction[0]
    mb2_history = correction[1]
    correct_prevq = ShowDataAdj(prevq_correction, comb_val)

    # main dataframes used
    whole_df = helper.RawDataCleanPrep().split_by_snap()

    kprev_df = whole_df[0]
    kact_df = whole_df[1]

    mb2_presence = fix_mb2_moves_per_quarters(mb2_history, kact_df, comb_val)

    # class and raw content init
    kprev = visuals.FormTable(kprev_df, comb_val)
    kact = visuals.FormTable(kact_df, comb_val)

    # for the main data sheet
    mdws = wb.add_worksheet("Main Data")

    # adding quarter names in dataframe form
    data = [tot_period[0]]
    data1 = [tot_period[1]]
    dfp = pd.DataFrame(data, columns=["Q List"])
    dfa = pd.DataFrame(data1, columns=["Q List"])

    dfp.to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=0,
        startrow=3,
        index=False,
        header=False,
    )
    dfa.to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=2,
        startrow=3,
        index=False,
        header=False,
    )

    correct_prevq.get_mb1().to_excel(
        writer, sheet_name=mdws.get_name(), startcol=0, startrow=4, index=False
    )
    kact.gen_mb1().to_excel(
        writer, sheet_name=mdws.get_name(), startcol=2, startrow=4, index=False
    )

    correct_prevq.get_population_numbers(tot_period[0]).to_excel(
        writer, sheet_name=mdws.get_name(), startcol=5, startrow=3, index=False
    )
    kact.gen_corrected_population(mb2_presence, tot_period[1]).to_excel(
        writer, sheet_name=mdws.get_name(), startcol=6, startrow=3, index=False
    )

    correct_prevq.get_fem_srt(tot_period[0]).to_excel(
        writer, sheet_name=mdws.get_name(), startcol=5, startrow=10, index=False
    )
    kact.corrected_female_srt(mb2_presence, tot_period[1]).to_excel(
        writer, sheet_name=mdws.get_name(), startcol=6, startrow=10, index=False
    )

    kprev.gen_fem34(tot_period[0]).to_excel(
        writer, sheet_name=mdws.get_name(), startcol=5, startrow=16, index=False
    )

    kact.gen_fem34(tot_period[1]).to_excel(
        writer, sheet_name=mdws.get_name(), startcol=6, startrow=16, index=False
    )

    # row index to place the new dataframe when the mb-1 name list varies or is larger than 10
    r = max(kact.gen_mb1().shape[0] + 6, correct_prevq.get_mb1().shape[0] + 6, 23) + 2

    dfp.to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=0,
        startrow=r + 1,
        index=False,
        header=False,
    )

    dfa.to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=4,
        startrow=r + 1,
        index=False,
        header=False,
    )

    correct_prevq.get_mb1_nationality(kprev_df).to_excel(
        writer, sheet_name=mdws.get_name(), startcol=0, startrow=r + 2, index=False
    )
    kact.gen_rlft().to_excel(
        writer, sheet_name=mdws.get_name(), startcol=4, startrow=r + 2, index=False
    )

    # row is the distance between the rlft dataframe and cross industry dataframe
    row = max(
        correct_prevq.get_mb1_nationality(kprev_df).shape[0] + r + 5,
        kact.gen_rlft().shape[0] + r + 5,
    )

    dfp.to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=0,
        startrow=row + 1,
        index=False,
        header=False,
    )
    dfa.to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=0,
        startrow=row + 1 + correct_prevq.gen_cross_ind(kprev_df).shape[0] + 4,
        index=False,
        header=False,
    )

    correct_prevq.gen_cross_ind(kprev_df).to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=0,
        startrow=row + 2,
        index=False,
    )
    kact.gen_cross_ind().to_excel(
        writer,
        sheet_name=mdws.get_name(),
        startcol=0,
        startrow=row + correct_prevq.gen_cross_ind(kprev_df).shape[0] + 6,
        index=False,
    )

    # MB-2 name list
    worksheet2 = wb.add_worksheet("MB-2 Name List")
    summarize_mb2(mb2_history, kact_df, comb_val).to_excel(
        writer,
        sheet_name=worksheet2.get_name(),
        startcol=3,
        startrow=4,
        index=False,
        header=False,
    )
    mb2_presence.to_excel(
        writer,
        sheet_name=worksheet2.get_name(),
        startcol=1,
        startrow=5,
        index=False,
    )

    # ws5 data
    worksheet5 = wb.add_worksheet("GR34 Full Name")
    visuals.gen_g34df_last(
        comb_val,
    ).to_excel(
        writer, sheet_name=worksheet5.get_name(), startcol=1, startrow=3, index=False
    )

    # ws6 data

    worksheet6 = wb.add_worksheet("Movements")

    a_list = movements.CalculateMovements(mb2_history, kact_df, comb_val).fill_in_ci()[
        0
    ]

    r_list = movements.CalculateMovements(mb2_history, kact_df, comb_val).fill_in_ci()[
        1
    ]

    a_list.to_excel(
        writer, sheet_name=worksheet6.get_name(), startcol=1, startrow=5, index=False
    )
    r_list.to_excel(
        writer, sheet_name=worksheet6.get_name(), startcol=6, startrow=5, index=False
    )

    wb.close()
    writer.close()

    # adding details on each worksheet
    from openpyxl.styles import Alignment, Border, Side

    def set_border(ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    dest_path = os.path.join(os.getcwd(), fl_loc).replace("\\", "\\")

    try:
        workb = openpyxl.load_workbook(dest_path)

        # adding details to main data sheet
        worksheet1 = workb["Main Data"]
        worksheet1["A2"] = "MB-1 total population"
        worksheet1.merge_cells("A2:C2")
        cc = worksheet1["A2"]
        cc.alignment = Alignment(horizontal="center")

        set_border(worksheet1, "A4:A4")
        set_border(worksheet1, "C4:C4")
        set_border(worksheet1, "E4:G6")
        set_border(worksheet1, "E11:G13")
        set_border(worksheet1, "E17:G20")

        worksheet1["E2"] = "MB-2 Total Population"
        worksheet1.merge_cells("E2:G2")
        ccr = worksheet1["E2"]
        ccr.alignment = Alignment(horizontal="center")

        worksheet1["E5"] = "Total MB-2 population"
        worksheet1["E5"].alignment = Alignment(horizontal="center")

        worksheet1["E6"] = "Total MB-1/2 population"
        worksheet1["E6"].alignment = Alignment(horizontal="center")

        worksheet1["E9"] = "Female % in senior teams"
        worksheet1.merge_cells("E9:G9")
        worksheet1["E9"].alignment = Alignment(horizontal="center")

        worksheet1["E12"] = "MB-1/2 Female #"
        worksheet1["E12"].alignment = Alignment(horizontal="center")

        worksheet1["E13"] = "MB-1/2 Female %"
        worksheet1["E13"].alignment = Alignment(horizontal="center")

        worksheet1["E15"] = "Female % G34+ incl. GGs"
        worksheet1.merge_cells("E15:G15")
        worksheet1["E15"].alignment = Alignment(horizontal="center")

        worksheet1["E18"] = "Total Population"
        worksheet1["E18"].alignment = Alignment(horizontal="center")

        worksheet1["E19"] = "Female #"
        worksheet1["E19"].alignment = Alignment(horizontal="center")

        worksheet1["E20"] = "Female %"
        worksheet1["E20"].alignment = Alignment(horizontal="center")

        kact_shape = kact.gen_mb1().shape[0] + 4
        kprev_shape = correct_prevq.get_mb1().shape[0] + 4

        row_rlft_ref = max(r, kact_shape, kprev_shape)

        worksheet1[f"A{row_rlft_ref}"] = "50% distinct nationalities in RLT/FLT"
        worksheet1.merge_cells(f"A{row_rlft_ref}:G{row_rlft_ref}")
        worksheet1[f"A{row_rlft_ref}"].alignment = Alignment(horizontal="center")

        title_cell_value = row if row_rlft_ref <= 22 else row
        title_cell = f"A{title_cell_value}"
        title_cell_range = f"A{title_cell_value}:G{title_cell_value}"

        worksheet1[title_cell] = "Cross Industry hires in senior teams past 3Y"
        worksheet1.merge_cells(title_cell_range)
        worksheet1[title_cell].alignment = Alignment(horizontal="center")
        try:
            worksheet1[f"C{row+2}"] = visuals.count_names(
                correct_prevq.gen_cross_ind(kprev_df)
            )
            worksheet1[f"C{row+2}"].alignment = Alignment(horizontal="center")
        except AttributeError:
            worksheet1[f"C{row+2}"] = visuals.count_names(
                correct_prevq.gen_cross_ind(kprev_df)
            )
            worksheet1[f"C{row+2}"].alignment = Alignment(horizontal="center")

        try:
            perc = round(
                (
                    visuals.count_names(correct_prevq.gen_cross_ind(kprev_df))
                    / (
                        visuals.count_names(correct_prevq.get_mb1())
                        + visuals.count_names(correct_prevq.get_mb2())
                    )
                    * 100
                ),
            )
            worksheet1[f"D{row+2}"] = f"{perc}%"
            worksheet1[f"D{row+2}"].alignment = Alignment(horizontal="center")
        except ZeroDivisionError:
            perc = 0
            worksheet1[f"D{row+2}"] = f"{perc}%"
            worksheet1[f"D{row+2}"].alignment = Alignment(horizontal="center")

        worksheet1[f"C{row + correct_prevq.gen_cross_ind(kprev_df).shape[0] +6}"] = len(
            kact.gen_cross_ind()
        )

        worksheet1[
            f"C{row + correct_prevq.gen_cross_ind(kprev_df).shape[0] +6}"
        ].alignment = Alignment(horizontal="center")

        try:
            ci_pos = row + correct_prevq.gen_cross_ind(kprev_df).shape[0] + 6
            ci_proc = round(
                (
                    visuals.count_names(kact.gen_cross_ind())
                    / (
                        visuals.count_names(kact.show_mb_1())
                        + visuals.count_names(kact.get_mb2_population())
                    )
                    * 100
                )
            )
        except ZeroDivisionError:
            ci_proc = "0%"

        worksheet1[f"D{ci_pos}"] = f"{ci_proc}%"
        worksheet1[f"D{ci_pos}"].alignment = Alignment(horizontal="center")

        # adding details to MB-2 Name List sheet
        try:
            worksheet2 = workb["MB-2 Name List"]
            worksheet2["B3"] = "MB-2 Name List"
            worksheet2.merge_cells("B3:K3")
            worksheet2["B3"].alignment = Alignment(horizontal="center")

            start_row = 6
            start_column = 2  # for column B
            col_index = int(mb2_presence.shape[1])
            end_row = start_row + len(mb2_presence)
            end_col = start_column + len(mb2_presence.columns)
            rnum = 6 + mb2_presence.shape[0]

            start_col_letter = get_excel_column_letter(start_column)
            end_col_letter = get_excel_column_letter(end_col)

            start_cell = f"{start_col_letter}{start_row}"
            end_cell = f"{end_col_letter}{end_row}"

            set_border(
                worksheet2,
                f"{start_cell}:{end_cell}",
            )
        except Exception:
            l.form_log(Exception, utils.get_level("Warn"))

        worksheet5 = workb["GR34 Full Name"]
        worksheet5["B2"] = "GR34+ Name List"
        worksheet5["B2"].alignment = Alignment(horizontal="center")

        set_border(
            worksheet5,
            f"B4:AH{4+visuals.gen_g34df_last(comb_val).shape[0]}",
        )

        worksheet6 = workb["Movements"]

        worksheet6["B3"] = "Movement List"
        worksheet6["B3"].alignment = Alignment(horizontal="center")

        workb.save(dest_path)

    except IOError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Main Data"
        wb.save(dest_path)
    l.form_log(f"{fl_loc} has been generated", 20)
    return dest_path


def add_details(file_loc):  # sourcery skip: low-code-quality
    with xw.App() as writer:
        workb = writer.books.open(file_loc)
        # workb.app.screen_updating = False

        workb.sheets["Main Data"].autofit()
        # will autofit columns and rows in sheet. To customize: pass 'c' for columns and 'r' for rows
        worksheet1 = workb.sheets["Main Data"]
        visuals.header_text_look(worksheet1, "A2")
        visuals.header_text_look(worksheet1, "E2")
        visuals.header_text_look(worksheet1, "E9")
        visuals.header_text_look(worksheet1, "E15")
        visuals.header2_text_look(worksheet1, "A4")
        visuals.header2_text_look(worksheet1, "C4")
        visuals.header2_text_look(worksheet1, "F4")
        visuals.header2_text_look(worksheet1, "G4")
        visuals.header2_text_look(worksheet1, "F11")
        visuals.header2_text_look(worksheet1, "G11")
        visuals.header2_text_look(worksheet1, "F17")
        visuals.header2_text_look(worksheet1, "G17")
        visuals.header3_text_look(worksheet1, "A5")
        visuals.header3_text_look(worksheet1, "C5")

        rlftpos = []
        crossind_pos = []
        for row, col in itertools.product(range(1, 100), range(1, 7)):
            if (
                worksheet1.range((row, col)).value
                == "50% distinct nationalities in RLT/FLT"
            ):
                rlftpos.extend((col, row))
            if (
                worksheet1.range((row, col)).value
                == "Cross Industry hires in senior teams past 3Y"
            ):
                crossind_pos.extend((col, row))

        visuals.adaptive_header1_style(worksheet1, rlftpos)
        visuals.adaptive_header1_style(worksheet1, crossind_pos)

        visuals.adaptive_header2_style(worksheet1, [rlftpos[0], rlftpos[1] + 2])
        visuals.adaptive_header2_style(worksheet1, [rlftpos[0] + 4, rlftpos[1] + 2])

        visuals.adaptive_q_cell_style(worksheet1, rlftpos)
        visuals.adaptive_header3_style(worksheet1, [rlftpos[0] + 4, rlftpos[1] + 3])
        visuals.adaptive_header3_style(worksheet1, [rlftpos[0] + 5, rlftpos[1] + 3])
        visuals.adaptive_header3_style(worksheet1, [rlftpos[0] + 6, rlftpos[1] + 3])

        visuals.adaptive_header2_style(
            worksheet1, [crossind_pos[0], crossind_pos[1] + 2]
        )
        visuals.ci_count_style(worksheet1, [crossind_pos[0] + 2, crossind_pos[1] + 2])

        visuals.adaptive_cross_industry_headers(worksheet1, crossind_pos, 3)

        ciq_cell = []
        for r, c in itertools.product(range(crossind_pos[1] + 3, 200), range(1, 7)):
            if re.match("^Q[1-4].[1-5]{2}", str(worksheet1.range((r, c)).value)):
                ciq_cell.extend((c, r))

        visuals.adaptive_header2_style(worksheet1, ciq_cell)
        visuals.ci_count_style(worksheet1, [ciq_cell[0] + 2, ciq_cell[1]])

        visuals.adaptive_cross_industry_headers(worksheet1, ciq_cell, 1)

        # worksheet2
        worksheet2 = workb.sheets["MB-2 Name List"]
        workb.sheets["MB-2 Name List"].autofit()
        worksheet2["B3"].color = (255, 192, 0)
        worksheet2.range("B3:B3").api.Font.Bold = True
        worksheet2.range("B3:B3").api.Font.ColorIndex = 0

        start_row = 5
        start_column = 4  # for column D
        end_column = start_column
        while True:
            cell_value = worksheet2.range(start_row, end_column + 1).value
            if cell_value is None:
                break
            end_column += 1

        start_col_letter = get_excel_column_letter(start_column)
        end_col_letter = get_excel_column_letter(end_column)

        start_cell = f"{start_col_letter}{start_row}"
        end_cell = f"{end_col_letter}{start_row}"
        cellrg = f"{start_cell}:{end_cell}"
        worksheet2.range(cellrg).api.Font.Bold = True
        worksheet2.range(cellrg).api.Font.ColorIndex = 25

        target_header = ["FullName", "Full Name", "Name"]
        mb2dfpos = None

        for col in range(1, 45):  # Assuming a maximum of 44 columns
            for row in range(1, 11):  # Assuming a maximum of 10 rows
                cell_value = worksheet2.range((row, col)).value
                if cell_value in target_header:
                    mb2dfpos = (col, row)
                    break
            if mb2dfpos is not None:
                break

        if not mb2dfpos:
            raise ValueError(
                f"The header '{target_header}' was not found in the specified range."
            )

        start_col, row = mb2dfpos
        col = start_col
        data_in_row = []
        while True:
            cell_value = worksheet2.range((row, col)).value
            if cell_value is None:
                break

            data_in_row.append(f"{get_excel_column_letter(col)}{row}")
            col += 1
        if not data_in_row:
            raise ValueError("No data found in the row.")

        final_range = ":".join(data_in_row)
        worksheet2[final_range].color = (0, 32, 96)
        worksheet2.range(final_range).api.Font.Bold = True
        worksheet2.range(final_range).api.Font.ColorIndex = 2

        worksheet2.range("B3:B3").api.Font.Bold = True
        worksheet2.range("B3:B3").api.Font.ColorIndex = 25

        # worksheet6
        worksheet6 = workb.sheets["Movements"]
        workb.sheets["Movements"].autofit()

        worksheet6["B3"].color = (255, 192, 0)
        worksheet6.range("B3:B3").api.Font.Bold = True
        worksheet6.range("B3:B3").api.Font.ColorIndex = 25

        worksheet6["B6:E6"].color = (0, 32, 96)
        worksheet6.range("B6:E6").api.Font.Bold = True
        worksheet6.range("B6:E6").api.Font.ColorIndex = 2

        worksheet6["G6:J6"].color = (0, 32, 96)
        worksheet6.range("G6:J6").api.Font.Bold = True
        worksheet6.range("G6:J6").api.Font.ColorIndex = 2

        workb.sheets["GR34 Full Name"].autofit()
        worksheet5 = workb.sheets["GR34 Full Name"]

        worksheet5["B2:B2"].color = (255, 192, 0)
        worksheet5.range("B2:B2").api.Font.Bold = True
        worksheet5.range("B2:B2").api.Font.ColorIndex = 25

        worksheet5["B4:AG4"].color = (47, 117, 181)
        worksheet5.range("B4:AG4").api.Font.Bold = True
        worksheet5.range("B4:AG4").api.Font.ColorIndex = 2

        workb.save(file_loc)
    return file_loc
